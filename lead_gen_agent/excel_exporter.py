"""Export leads to a formatted Excel workbook."""
from __future__ import annotations
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from .models import Lead, LeadType, ICP

# Colour palette
_HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
_HOT_FILL      = PatternFill("solid", fgColor="FFE0E0")   # light red
_WARM_FILL     = PatternFill("solid", fgColor="FFF8DC")   # light yellow
_COLD_FILL     = PatternFill("solid", fgColor="E0EFFF")   # light blue
_ALT_FILL      = PatternFill("solid", fgColor="F5F5F5")   # subtle grey

_HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT     = Font(name="Calibri", size=10)
_TITLE_FONT    = Font(name="Calibri", bold=True, size=14, color="1F3864")

_THIN_BORDER   = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

_COLUMNS = [
    ("Rank",         8),
    ("Company Name", 24),
    ("Website",      30),
    ("Industry",     20),
    ("Location",     22),
    ("Description",  42),
    ("Contact Email",28),
    ("Lead Score",   12),
    ("Fit %",        10),
    ("Lead Type",    12),
    ("Reason",       55),
]

_TYPE_FILL = {
    LeadType.HOT:  _HOT_FILL,
    LeadType.WARM: _WARM_FILL,
    LeadType.COLD: _COLD_FILL,
}

_TYPE_COLOR = {
    LeadType.HOT:  "C0392B",
    LeadType.WARM: "D35400",
    LeadType.COLD: "2471A3",
}


def _apply_header(ws, row: int) -> None:
    for col, (label, _) in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _set_col_widths(ws) -> None:
    for col, (_, width) in enumerate(_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def export_to_excel(
    leads: List[Lead],
    icp: ICP,
    business_description: str,
    out_path: Path,
) -> None:
    wb = Workbook()

    # ── Sheet 1: Leads ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Leads"
    ws.freeze_panes = "A4"

    # Title row
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"Lead Generation Results — {business_description[:80]}"
    title_cell.font  = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # Blank spacer
    ws.row_dimensions[2].height = 6

    # Header
    _apply_header(ws, row=3)
    ws.row_dimensions[3].height = 22

    # Data rows
    for rank, lead in enumerate(leads, 1):
        row = rank + 3
        fill = _TYPE_FILL.get(lead.lead_type, _ALT_FILL)
        color = _TYPE_COLOR.get(lead.lead_type, "000000")

        values = [
            rank,
            lead.company_name,
            lead.website,
            lead.industry,
            lead.location or "Unknown",
            lead.description,
            lead.contact_email or "—",
            lead.lead_score,
            lead.fit_percentage,
            lead.lead_type.value if hasattr(lead.lead_type, "value") else lead.lead_type,
            lead.reason,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill   = fill
            cell.border = _THIN_BORDER
            cell.font   = Font(
                name="Calibri", size=10,
                bold=(col == 7),          # bold score column
                color=(color if col == 9 else "000000"),
            )
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
                horizontal="center" if col in (1, 7, 8, 9) else "left",
            )

        ws.row_dimensions[row].height = 48

    _set_col_widths(ws)

    # ── Sheet 2: ICP Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("ICP Summary")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 70

    icp_rows = [
        ("Business Description", business_description),
        ("Target Industries",    ", ".join(icp.target_industries)),
        ("Company Size",         icp.company_size),
        ("Geographies",          ", ".join(icp.geographies)),
        ("Pain Points",          "\n".join(f"• {p}" for p in icp.pain_points)),
        ("Keywords",             ", ".join(icp.keywords)),
        ("ICP Summary",          icp.description),
    ]

    for r, (label, value) in enumerate(icp_rows, 1):
        a = ws2.cell(row=r, column=1, value=label)
        a.font = Font(name="Calibri", bold=True, size=10, color="1F3864")
        a.alignment = Alignment(vertical="top")
        a.fill = PatternFill("solid", fgColor="EBF2FF")
        a.border = _THIN_BORDER

        b = ws2.cell(row=r, column=2, value=value)
        b.font = Font(name="Calibri", size=10)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        b.border = _THIN_BORDER
        ws2.row_dimensions[r].height = max(40, value.count("\n") * 16 + 20)

    wb.save(out_path)
